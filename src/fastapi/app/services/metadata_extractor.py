import os
import hashlib
import magic
import fitz  # pymupdf
import openpyxl
from pathlib import Path
from docx import Document

def compute_sha256(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()

def extract_full_text_pdf(path: str) -> str:
    try:
        doc = fitz.open(path)
        text = " ".join(page.get_text() for page in doc)
        doc.close()
        return text.strip()
    except Exception:
        return ""

def extract_full_text_txt(path: str) -> str:
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read().strip()
    except Exception:
        return ""

def extract_full_text_docx(path: str) -> str:
    try:
        doc = Document(path)
        return " ".join(p.text for p in doc.paragraphs if p.text.strip()).strip()
    except Exception:
        return ""

def extract_full_text_xlsx(path: str) -> str:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        parts = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    text = str(cell).strip()
                    if text:
                        parts.append(text)
        wb.close()
        return " ".join(parts).strip()
    except Exception:
        return ""

def extract_metadata(tmp_path: str, storage_path: str) -> dict:
    ext = Path(storage_path).suffix.lower()
    file_size_kb = round(os.path.getsize(tmp_path) / 1024, 2)
    mime_type = magic.from_file(tmp_path, mime=True)
    hash_sha256 = compute_sha256(tmp_path)

    extra = {}
    full_text = ""

    # PDF
    if ext == ".pdf":
        try:
            doc = fitz.open(tmp_path)
            extra["page_count"] = doc.page_count
            doc.close()
        except Exception:
            extra["page_count"] = None
        full_text = extract_full_text_pdf(tmp_path)

    # Excel
    elif ext in (".xlsx", ".xlsm"):
        try:
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            extra["sheet_count"] = len(wb.sheetnames)
            extra["sheet_names"] = wb.sheetnames
            wb.close()
        except Exception:
            extra["sheet_count"] = None
        full_text = extract_full_text_xlsx(tmp_path)

    # Texto plano / CSV / XML
    elif ext in (".txt", ".csv", ".xml", ".json"):
        try:
            with open(tmp_path, "r", encoding="utf-8", errors="ignore") as f:
                lines = f.readlines()
            extra["line_count"] = len(lines)
        except Exception:
            extra["line_count"] = None
        full_text = extract_full_text_txt(tmp_path)

    # Word
    elif ext in (".docx",):
        extra["note"] = "Documento Word"
        full_text = extract_full_text_docx(tmp_path)

    # CAD / binarios
    elif ext in (".dwg", ".dxf", ".stl", ".step", ".igs"):
        extra["note"] = "Archivo CAD/binario, sin extracción de contenido"

    else:
        extra["note"] = f"Tipo no manejado específicamente: {ext}"

    return {
        "extension": ext,
        "mime_type": mime_type,
        "file_size_kb": file_size_kb,
        "hash_sha256": hash_sha256,
        "full_text": full_text,
        **extra
    }